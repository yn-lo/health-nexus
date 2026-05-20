import openpyxl
import re
from datetime import date, datetime
from typing import List, Tuple
from django.contrib.auth import get_user_model
from django.db import transaction
from apps.care.models import PatientProfile
from apps.base.models import Department

User = get_user_model()


class BulkImportResult:
    def __init__(self):
        self.success_count = 0
        self.fail_count = 0
        self.errors: List[str] = []


class BulkImportService:
    """用户批量导入服务（支持患者/医生/护士）"""

    REQUIRED_HEADERS = ['姓名', '手机号', '科室', '性别', '角色']
    OPTIONAL_HEADERS = ['出生日期']

    def validate_headers(self, headers: List[str]) -> Tuple[bool, str]:
        for req in self.REQUIRED_HEADERS:
            if req not in headers:
                return False, f"缺少必需列: {req}"
        return True, ""

    def import_from_excel(self, file, creator: User) -> BulkImportResult:
        result = BulkImportResult()
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        valid, error = self.validate_headers(headers)
        if not valid:
            result.errors.append(error)
            result.fail_count = ws.max_row - 1
            return result

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            try:
                data = dict(zip(headers, row))
                self._create_user(data, creator)
                result.success_count += 1
            except Exception as e:
                result.fail_count += 1
                result.errors.append(f"第{row_idx}行: {str(e)}")

        return result

    @transaction.atomic
    def _create_user(self, data: dict, creator: User):
        name = data.get('姓名')
        phone = str(data.get('手机号', '')).strip()
        dept_name = data.get('科室', '').strip()
        gender = data.get('性别', '').strip().upper()
        birth_date = data.get('出生日期')
        role_name = (data.get('角色') or '').strip()

        if not name or not phone:
            raise ValueError("姓名和手机号不能为空")

        if not role_name:
            raise ValueError("角色不能为空")

        role_map = {'患者': User.Role.PATIENT, '医生': User.Role.DOCTOR, '护士': User.Role.NURSE}
        role = role_map.get(role_name)
        if not role:
            raise ValueError(f"角色无效: {role_name}（支持: 患者/医生/护士）")

        gender_map = {'M': 'M', 'F': 'F', 'O': 'O', '男': 'M', '女': 'F', '其他': 'O'}
        gender = gender_map.get(gender, 'O')

        birth_date = self._parse_birth_date(data.get('出生日期'))

        dept = Department.objects.filter(name=dept_name).first()
        if not dept:
            raise ValueError(f"科室不存在: {dept_name}")

        user = User.objects.create_user(
            username=phone,
            phone=phone,
            role=role,
        )

        if role == User.Role.PATIENT:
            PatientProfile.objects.create(
                user=user,
                name=name,
                gender=gender,
                birth_date=birth_date,
            )
            user.patient_profile.departments.add(dept)
        else:
            user.departments.add(dept)

        return user

    @staticmethod
    def _parse_birth_date(value) -> date:
        if value is None:
            return None

        if isinstance(value, (date, datetime)):
            return value if isinstance(value, date) else value.date()

        if isinstance(value, (int, float)):
            try:
                from datetime import timedelta
                base_date = date(1899, 12, 30)
                return base_date + timedelta(days=int(value))
            except (ValueError, OverflowError):
                return None

        if isinstance(value, str):
            value = value.strip()
            for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日', '%Y.%m.%d'):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue

        return None
